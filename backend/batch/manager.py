"""批量任务管理器 — asyncio.Queue + BackgroundTasks.

处理 Excel 批量排查，支持暂停/继续/取消。
每个任务独立协程运行，通过 WebSocket 推送进度。
v2: 支持内部并发（Semaphore 控制） + 流式 Excel 写入。
"""

from __future__ import annotations

import asyncio
import os
import tempfile
import uuid
from io import BytesIO
from pathlib import Path

import openpyxl

from backend.db.database import get_connection
from backend.logging_setup import get_logger
from backend.models.schemas import FetchResult

logger = get_logger()


class BatchTask:
    """单个批量任务."""

    def __init__(
        self,
        task_id: str,
        filename: str,
        packages: list[tuple[str, str | None, str | None]],
        excel_bytes: bytes = b"",
        pkg_col: int = 0,
        evc_col: int = 0,
        ev_col: int = 0,
    ):
        self.id = task_id
        self.filename = filename
        self.packages = packages
        self.total = len(packages)
        self.completed = 0
        self.status = "pending"
        self.results: list[FetchResult] = []
        self.stop_event = asyncio.Event()
        self._progress_callbacks: list[callable] = []
        # 原始 Excel 信息 (用于完成后在原文件写入结果)
        self.excel_bytes = excel_bytes
        self.pkg_col = pkg_col
        self.evc_col = evc_col
        self.ev_col = ev_col
        # 临时结果文件路径（流式写入，避免内存堆积）
        self._temp_result_path: str = ""

    def on_progress(self, callback):
        self._progress_callbacks.append(callback)

    async def notify_progress(self):
        for cb in self._progress_callbacks:
            try:
                await cb(self)
            except Exception:
                pass

    def to_dict(self) -> dict:
        # 统计结果
        matched = newer = older = not_found = 0
        for r in self.results:
            s = r.compare_status.value
            if s == "matched": matched += 1
            elif s == "newer": newer += 1
            elif s == "older": older += 1
            else: not_found += 1

        return {
            "id": self.id,
            "filename": self.filename,
            "total": self.total,
            "completed": self.completed,
            "status": self.status,
            "progress_pct": round(self.completed / self.total * 100, 1) if self.total else 0,
            "summary": {
                "matched": matched,
                "newer": newer,
                "older": older,
                "not_found": not_found,
            },
        }


class BatchManager:
    """批量任务管理器.

    使用 asyncio.Queue 接收任务，独立协程处理，
    支持暂停/继续/取消。
    """

    def __init__(self):
        self._queue: asyncio.Queue[BatchTask] = asyncio.Queue()
        self._active_tasks: dict[str, BatchTask] = {}

    async def enqueue(self, task: BatchTask) -> None:
        """添加批量任务."""
        await self._queue.put(task)
        self._active_tasks[task.id] = task
        self._save_task(task)
        logger.info("批量任务已入队: {} ({} 行)", task.filename, task.total)

    async def run(self, task: BatchTask):
        """执行批量任务（并发处理多个包名，默认5并发）.

        使用 asyncio.Semaphore 控制并发包名数，
        每个包名内仍按默认站点并发（scraper_concurrency）查询。
        """
        from backend.config import get_settings
        from backend.core.orchestrator import query_fast

        task.status = "running"
        self._save_task(task)
        await task.notify_progress()

        settings = get_settings()
        batch_concurrency = getattr(settings, "batch_concurrency", 5)
        semaphore = asyncio.Semaphore(batch_concurrency)

        async def _process_one(idx: int, pkg: str, ev: str | None, evc: str | None):
            async with semaphore:
                # 检查取消
                if task.status == "cancelled":
                    return idx, None

                # 等待暂停恢复
                while task.status == "paused":
                    await asyncio.sleep(0.5)

                if task.status == "cancelled":
                    return idx, None

                try:
                    result = await query_fast(pkg, ev, evc)
                except Exception as e:
                    logger.warning("批量任务行 {} 失败: {} — {}", idx + 1, pkg, e)
                    from backend.models.schemas import CompareStatus
                    result = FetchResult(
                        package=pkg,
                        error=str(e),
                        compare_status=CompareStatus.ERROR,
                    )
                return idx, result

        # 创建所有协程（由 Semaphore 限流）
        coros = [_process_one(i, pkg, ev, evc) for i, (pkg, ev, evc) in enumerate(task.packages)]

        # 使用 as_completed 实时跟踪进度
        results_dict: dict[int, FetchResult] = {}
        for coro in asyncio.as_completed(coros):
            if task.status == "cancelled":
                break
            idx, result = await coro
            if result is not None:
                results_dict[idx] = result
                task.completed = len(results_dict)
                # 进度推送（控制频率：每5个推送一次，或第一个/最后一个）
                if task.completed % 5 == 0 or task.completed == 1:
                    self._save_task(task)
                    await task.notify_progress()

        # 按原始顺序排列结果
        if task.status != "cancelled":
            ordered = [results_dict[i] for i in sorted(results_dict.keys()) if i in results_dict]
            task.results = ordered
            task.status = "completed"
        else:
            # 取消后也保留已完成的结果
            ordered = [results_dict[i] for i in sorted(results_dict.keys()) if i in results_dict]
            task.results = ordered
        task.completed = len(task.results)

        # 流式持久化结果到临时文件，释放内存
        if ordered:
            self._flush_results_to_temp(task, ordered)

        self._save_task(task)
        await task.notify_progress()
        logger.info("批量任务完成: {} ({}/{})", task.filename, task.completed, task.total)

    def _flush_results_to_temp(self, task: BatchTask, results: list[FetchResult]) -> None:
        """将结果流式写入临时 JSONL 文件，降低内存占用."""
        import json
        try:
            if not task._temp_result_path:
                fd, task._temp_result_path = tempfile.mkstemp(
                    suffix=".jsonl", prefix=f"batch_{task.id}_"
                )
                os.close(fd)
            with open(task._temp_result_path, "w", encoding="utf-8") as f:
                for r in results:
                    f.write(json.dumps(r.to_dict(), ensure_ascii=False) + "\n")
            logger.debug("结果已流式写入: {} ({} 条)", task._temp_result_path, len(results))
        except Exception as e:
            logger.warning("流式写入临时文件失败: {}", e)

    @staticmethod
    def _load_results_from_temp(task: "BatchTask") -> list[dict]:
        """从临时 JSONL 文件加载结果."""
        import json
        results = []
        if task._temp_result_path and os.path.exists(task._temp_result_path):
            with open(task._temp_result_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line:
                        results.append(json.loads(line))
        return results

    @staticmethod
    def _cleanup_temp(task: "BatchTask") -> None:
        """清理临时结果文件."""
        if task._temp_result_path and os.path.exists(task._temp_result_path):
            try:
                os.unlink(task._temp_result_path)
                logger.debug("临时文件已清理: {}", task._temp_result_path)
            except Exception as e:
                logger.warning("清理临时文件失败: {}", e)

    def pause(self, task_id: str) -> bool:
        if task_id in self._active_tasks:
            task = self._active_tasks[task_id]
            task.status = "paused"
            task.stop_event.set()
            return True
        return False

    def resume(self, task_id: str) -> bool:
        if task_id in self._active_tasks:
            task = self._active_tasks[task_id]
            if task.status == "paused":
                task.status = "running"
                task.stop_event.clear()
            return True
        return False

    def cancel(self, task_id: str) -> bool:
        if task_id in self._active_tasks:
            task = self._active_tasks[task_id]
            task.status = "cancelled"
            task.stop_event.set()
            BatchManager._cleanup_temp(task)
            return True
        return False

    def get_task(self, task_id: str) -> BatchTask | None:
        return self._active_tasks.get(task_id)

    def get_all_tasks(self) -> list[BatchTask]:
        return list(self._active_tasks.values())

    def _save_task(self, task: BatchTask):
        conn = get_connection()
        conn.execute("""
            INSERT OR REPLACE INTO batch_tasks
            (id, filename, total_rows, completed_rows, status, updated_at)
            VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """, (task.id, task.filename, task.total, task.completed, task.status))
        conn.commit()

    @staticmethod
    def export_to_excel(task: BatchTask) -> BytesIO:
        """在原 Excel 写入排查结果 (排查时间 + 版本号 + 对比状态).

        逻辑:
          1. 加载原始 Excel（保留样式）
          2. 在包名列右侧插入三列: 排查时间, 当前后台版本号(vc), 对比状态
          3. 若已有列则复用, 否则新建
          4. 返回修改后的文件
        """
        from datetime import datetime
        from openpyxl.styles import Alignment, Font, PatternFill

        HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        HEADER_FONT = Font(color="FFFFFF", bold=True)
        CENTER = Alignment(horizontal="center")

        wb = openpyxl.load_workbook(BytesIO(task.excel_bytes))
        ws = wb.active

        pkg_col = task.pkg_col
        if pkg_col <= 0:
            for c in range(1, ws.max_column + 1):
                h = str(ws.cell(1, c).value or "").lower().replace(" ", "").replace("_", "")
                if h in ("packagename", "package", "pkg"):
                    pkg_col = c
                    break
            if pkg_col <= 0:
                pkg_col = 1

        # 三列配置: (列名, 关键词, 列宽, 取值函数)
        columns = [
            ("排查时间", ["排查时间", "checktime", "check_time"], 18,
             lambda r, pkg: datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("当前后台版本号(vc)", ["当前后台版本号", "版本号(vc)", "versioncode", "version_code", "vc"], 18,
             lambda r, pkg: r.best_version_code or "-"),
            ("对比状态", ["对比状态", "状态", "status"], 14,
             lambda r, pkg: {"matched": "已匹配", "newer": "有新版本", "older": "版本较旧",
                              "not_found": "未找到", "error": "错误"}.get(r.compare_status.value, r.compare_status.value) if r else "-"),
        ]

        # 构建结果映射（优先用内存中的 results，否则从临时文件加载）
        result_map: dict[str, FetchResult] = {}
        if task.results:
            for r in task.results:
                result_map[r.package] = r
        elif task._temp_result_path:
            loaded = BatchManager._load_results_from_temp(task)
            for d in loaded:
                # 从 dict 重建简单的包装对象
                pkg = d.get("package", "")
                result_map[pkg] = d  # type: ignore

        # 查找或插入列
        col_positions = []
        for col_name, keywords, width, _ in columns:
            found = 0
            for c in range(1, ws.max_column + 1):
                h = str(ws.cell(1, c).value or "").lower().replace(" ", "").replace("_", "")
                for kw in keywords:
                    if kw.replace(" ", "").replace("_", "") in h:
                        found = c
                        break
                if found:
                    break
            if not found:
                # 在现有列右侧插入
                insert_at = ws.max_column + 1
                ws.insert_cols(insert_at)
                found = insert_at
                cell = ws.cell(1, found, col_name)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER
            col_positions.append(found)
            ws.column_dimensions[openpyxl.utils.get_column_letter(found)].width = width

        # 写入数据
        filled = 0
        for row in range(2, ws.max_row + 1):
            pkg = str(ws.cell(row, pkg_col).value or "").strip()
            if not pkg:
                continue
            r = result_map.get(pkg)
            for ci, (_, _, _, getter) in enumerate(columns):
                if isinstance(r, dict):
                    # 从临时文件加载的 dict 格式，手动取值
                    if ci == 0:
                        val = datetime.now().strftime("%Y-%m-%d %H:%M")
                    elif ci == 1:
                        val = str(r.get("best_version_code", "-")) or "-"
                    else:
                        s = r.get("compare_status", "")
                        val = {"matched": "已匹配", "newer": "有新版本", "older": "版本较旧",
                               "not_found": "未找到", "error": "错误"}.get(s, s)
                else:
                    val = getter(r, pkg) if r else (getter(None, pkg) if ci == 0 else "-")
                ws.cell(row, col_positions[ci], val).alignment = CENTER
            if r and not isinstance(r, dict) and r.best_version_code and r.best_version_code != "-":
                filled += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


# 全局单例
_batch_manager: BatchManager | None = None


def get_batch_manager() -> BatchManager:
    global _batch_manager
    if _batch_manager is None:
        _batch_manager = BatchManager()
    return _batch_manager
