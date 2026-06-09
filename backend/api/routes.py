"""REST API 路由 — 所有端点定义."""

from __future__ import annotations

import asyncio
import uuid
from io import BytesIO
from pathlib import Path
from typing import Any

from email.utils import parsedate_to_datetime, format_datetime

from fastapi import APIRouter, HTTPException, UploadFile, File, WebSocket, WebSocketDisconnect, Request, Query
from fastapi.responses import StreamingResponse, JSONResponse, Response
from slowapi import Limiter
from slowapi.util import get_remote_address
import openpyxl

from backend.api.websocket import get_ws_manager
from backend.batch.manager import BatchTask, BatchManager, get_batch_manager
from backend.config import get_settings, reload_settings
from backend.core.orchestrator import query_single, query_fast, query_slow, query_batch
from backend.core.cache import get_scraper_cache, get_slow_task_store
from backend.db.database import init_db
from backend.download.extractors import extract_download_links, pick_best_variant
from backend.download.manager import DownloadTask as DlTask, get_download_manager
from backend.logging_setup import get_logger
from backend.memo.store import get_memo_store

logger = get_logger()

router = APIRouter(prefix="/api")
limiter = Limiter(key_func=get_remote_address)


# ── 健康检查 ───────────────────────────────────────────────

@router.get("/health")
async def health():
    return {"status": "ok"}


@router.post("/test-proxy")
async def test_proxy():
    """测试代理连通性 — 通过代理访问 apkcombo.com 验证.

    返回: {"ok": true, "latency_ms": 123} 或 {"ok": false, "error": "..."}
    """
    import time
    from backend.config import get_settings
    from backend.core.http_client import _http_get_sync as http_get_sync

    settings = get_settings()
    if not settings.proxy:
        return {"ok": False, "error": "未配置代理"}

    test_url = "https://apkcombo.com"
    start = time.time()
    try:
        status, html = await asyncio.to_thread(http_get_sync, test_url)
        latency = round((time.time() - start) * 1000)
        if status == 200 and len(html) > 500:
            return {"ok": True, "latency_ms": latency}
        return {"ok": False, "error": f"HTTP {status}", "latency_ms": latency}
    except Exception as e:
        return {"ok": False, "error": str(e), "latency_ms": round((time.time() - start) * 1000)}


# ── 爬取 ───────────────────────────────────────────────────

# ── 爬取（三级模式）─────────────────────────────────────────

@router.post("/fetch")
@limiter.limit("30/minute")
async def fetch_package(request: Request, data: dict[str, Any]):
    """单包名查询（默认快速排查: Google Play + APKPure + APKCombo）."""
    return await _do_fetch(data, query_fast)


@router.post("/fetch/fast")
@limiter.limit("30/minute")
async def fetch_fast(request: Request, data: dict[str, Any]):
    """快速排查 — Google Play + APKPure + APKCombo（秒级响应）."""
    return await _do_fetch(data, query_fast)


@router.post("/fetch/slow")
@limiter.limit("10/minute")
async def fetch_slow(request: Request, data: dict[str, Any]):
    """慢速排查 — APKMirror + APKVision（浏览器渲染，30-90s）.

    同步阻塞模式，推荐使用 /api/fetch/slow/async 异步提交.
    """
    return await _do_fetch(data, query_slow)


# ── 慢速异步任务 ───────────────────────────────────────────

@router.post("/fetch/slow/async")
@limiter.limit("10/minute")
async def fetch_slow_async(request: Request, data: dict[str, Any]):
    """异步慢速排查 — 提交后台任务，立即返回 task_id.

    结果通过 GET /api/fetch/slow/result/{task_id} 轮询，
    或通过 WebSocket /api/ws/{task_id} 订阅完成通知。"""
    package = data.get("package", "").strip()
    if not package:
        raise HTTPException(400, "package is required")

    expected_version = data.get("expected_version")
    expected_version_code = data.get("expected_version_code")

    store = get_slow_task_store()
    task_id = store.create()

    async def _run_slow():
        try:
            result = await query_slow(package, expected_version, expected_version_code)
            store.complete(task_id, result)
            # WebSocket 推送完成通知
            ws_mgr = get_ws_manager()
            await ws_mgr.send_to_task(task_id, {
                "type": "slow_task_completed",
                "data": {"task_id": task_id, "status": "completed"},
            })
        except Exception as e:
            store.fail(task_id, str(e))
            ws_mgr = get_ws_manager()
            await ws_mgr.send_to_task(task_id, {
                "type": "slow_task_error",
                "data": {"task_id": task_id, "status": "error", "error": str(e)},
            })

    asyncio.create_task(_run_slow())

    return {
        "task_id": task_id,
        "status": "pending",
        "message": "任务已提交，预计 1-2 分钟完成",
    }


@router.get("/fetch/slow/result/{task_id}")
async def fetch_slow_result(task_id: str):
    """查询慢速异步任务结果."""
    store = get_slow_task_store()
    task = store.get(task_id)
    if not task:
        raise HTTPException(404, "任务不存在或已过期")
    if task["status"] == "pending":
        return {"task_id": task_id, "status": "pending", "result": None}
    if task["status"] == "error":
        return {"task_id": task_id, "status": "error", "error": task["error"]}
    if task["status"] == "completed" and task["result"]:
        return {
            "task_id": task_id,
            "status": "completed",
            "result": task["result"].to_dict(),
        }
    return {"task_id": task_id, "status": task["status"]}


@router.post("/fetch/all")
@limiter.limit("10/minute")
async def fetch_all(request: Request, data: dict[str, Any]):
    """全量排查 — 所有启用的站点（快速 + 慢速）."""
    return await _do_fetch(data, query_single)


async def _do_fetch(data: dict[str, Any], query_fn) -> dict:
    """统一处理单包名查询请求."""
    package = data.get("package", "").strip()
    if not package:
        raise HTTPException(400, "package is required")

    expected_version = data.get("expected_version")
    expected_version_code = data.get("expected_version_code")
    save_memo = data.get("save_memo", False)

    result = await query_fn(package, expected_version, expected_version_code)

    if save_memo and result.best_version:
        memo = get_memo_store()
        memo.upsert(
            package_name=package,
            version_code=result.best_version_code,
            version_name=result.best_version,
        )

    return result.to_dict()


@router.post("/fetch/batch")
@limiter.limit("10/minute")
async def fetch_batch(request: Request, data: dict[str, Any]):
    """多包名查询（使用内部并发控制，默认最大5并发）."""
    packages = data.get("packages", [])
    mode = data.get("mode", "fast")
    if not packages:
        raise HTTPException(400, "packages is required")

    # 解析包名列表
    parsed = []
    for item in packages:
        if isinstance(item, str):
            parsed.append((item, None, None))
        else:
            parsed.append((
                item.get("package", ""),
                item.get("expected_version"),
                item.get("expected_version_code"),
            ))

    pkg_list = [(pkg, ev, evc) for pkg, ev, evc in parsed if pkg]
    total = len(pkg_list)
    ws_mgr = get_ws_manager()

    async def _on_progress(completed: int, _total: int, _result):
        """通过全局 WebSocket 推送批量查询进度."""
        pct = round(completed / _total * 100, 1) if _total else 0
        await ws_mgr.broadcast({
            "type": "batch_fetch_progress",
            "data": {"completed": completed, "total": _total, "progress_pct": pct, "mode": mode},
        })

    # 使用 orchestrator.query_batch 进行限流并发查询
    results_list = await query_batch(
        pkg_list,
        mode=mode,
        progress_callback=_on_progress,
    )

    return {
        "results": [r.to_dict() for r in results_list],
        "mode": mode,
        "total": total,
    }


# ── 批量 Excel ─────────────────────────────────────────────

@router.post("/batch/upload")
async def batch_upload(
    file: UploadFile = File(...),
):
    """上传 Excel 文件并启动批量排查."""
    # 安全检查: 扩展名 + MIME + 大小限制
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "仅支持 .xlsx / .xls 文件")

    allowed_mimes = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/octet-stream",
    }
    if file.content_type and file.content_type not in allowed_mimes:
        raise HTTPException(400, f"不支持的文件类型: {file.content_type}")

    content = await file.read()
    max_size = 50 * 1024 * 1024  # 50MB
    if len(content) > max_size:
        raise HTTPException(400, f"文件过大 (最大 50MB, 当前 {len(content) // (1024*1024)}MB)")

    if len(content) < 128:
        raise HTTPException(400, "文件过小或损坏")
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    ws = wb.active

    # 自动检测列
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    pkg_col = None
    evc_col = None
    ev_col = None
    for i, h in enumerate(headers):
        h_lower = str(h).lower().replace(" ", "_") if h else ""
        if h_lower in ("package_name", "package", "packagename"):
            pkg_col = i
        elif h_lower in ("expected_version_code", "version_code", "versioncode"):
            evc_col = i
        elif h_lower in ("expected_version_name", "version_name", "versionname"):
            ev_col = i

    if pkg_col is None:
        raise HTTPException(400, "未找到 package_name 列")

    packages = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pkg = str(row[pkg_col]).strip() if row[pkg_col] else None
        if not pkg:
            continue
        ev = str(row[ev_col]).strip() if ev_col is not None and row[ev_col] else None
        evc = str(row[evc_col]).strip() if evc_col is not None and row[evc_col] else None
        packages.append((pkg, ev, evc))

    wb.close()

    if not packages:
        raise HTTPException(400, "Excel 中无有效数据")

    task_id = f"batch_{uuid.uuid4().hex[:8]}"
    batch_task = BatchTask(
        task_id=task_id,
        filename=file.filename,
        packages=packages,
        excel_bytes=content,
        pkg_col=pkg_col,
        evc_col=evc_col or 0,
        ev_col=ev_col or 0,
    )

    # WebSocket 进度回调
    async def on_progress(bt: BatchTask):
        ws_mgr = get_ws_manager()
        await ws_mgr.send_to_task(task_id, {
            "type": "batch_progress",
            "data": bt.to_dict(),
        })

    batch_task.on_progress(on_progress)

    manager = get_batch_manager()
    await manager.enqueue(batch_task)
    asyncio.create_task(manager.run(batch_task))

    return {
        "task_id": task_id,
        "total": batch_task.total,
        "filename": file.filename,
    }


@router.get("/batch/{task_id}")
async def batch_status(task_id: str):
    """查询批量任务状态."""
    manager = get_batch_manager()
    task = manager.get_task(task_id)
    if not task:
        raise HTTPException(404, "任务不存在")
    return task.to_dict()


@router.post("/batch/{task_id}/pause")
async def batch_pause(task_id: str):
    manager = get_batch_manager()
    if not manager.pause(task_id):
        raise HTTPException(404, "任务不存在")
    return {"ok": True}


@router.post("/batch/{task_id}/resume")
async def batch_resume(task_id: str):
    manager = get_batch_manager()
    if not manager.resume(task_id):
        raise HTTPException(404, "任务不存在")
    return {"ok": True}


@router.post("/batch/{task_id}/cancel")
async def batch_cancel(task_id: str):
    manager = get_batch_manager()
    if not manager.cancel(task_id):
        raise HTTPException(404, "任务不存在")
    return {"ok": True}


@router.get("/batch/{task_id}/download")
async def batch_download_result(task_id: str):
    """下载批量排查结果 Excel."""
    manager = get_batch_manager()
    task = manager.get_task(task_id)
    if not task:
        raise HTTPException(404, "任务不存在")
    if task.status != "completed":
        raise HTTPException(400, "任务未完成")

    output = BatchManager.export_to_excel(task)
    BatchManager._cleanup_temp(task)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=batch_{task_id}.xlsx"},
    )


# ── 路径下载安全 ───────────────────────────────────────────

import re as _re

_PKG_RE = _re.compile(r'^[a-zA-Z][a-zA-Z0-9_.]{1,127}$')

def _validate_package_name(pkg: str) -> str:
    """验证 Android 包名格式，防止路径遍历。

    Android 包名规则:
        - 以字母开头
        - 仅字母、数字、下划线和点号
        - 1-128 字符 (每个段 ≤127 字符)

    Raises:
        HTTPException: 包名无效。
    """
    pkg = pkg.strip()
    if not pkg or not _PKG_RE.match(pkg):
        raise HTTPException(400, f"包名格式无效: {pkg}")
    if ".." in pkg:
        raise HTTPException(400, "包名包含非法字符")
    return pkg


def _safe_save_path(settings, package: str, version: str, arch: str) -> str:
    """构造安全的保存路径，确保不超出 download_path。

    策略:
        1. 规范化 download_path 基础路径
        2. 清理 package/version 中的路径分隔符
        3. 验证最终路径在 download_path 子树内
    """
    base = Path(settings.download_path).resolve()
    safe_pkg = package.replace("\\", "/").replace("../", "").replace("..\\", "").strip("/")
    safe_ver = version.replace("\\", "/").replace("../", "").replace("..\\", "").strip("/")
    safe_arch = arch.replace("\\", "/").replace("../", "").replace("..\\", "").strip("/") or "unknown"

    target = (base / safe_pkg / safe_ver / f"{safe_pkg}_{safe_ver}_{safe_arch}.apk").resolve()
    try:
        target.relative_to(base)
    except ValueError:
        raise HTTPException(400, "非法保存路径")
    return str(target)


# ── 下载 ───────────────────────────────────────────────────

@router.post("/download")
@limiter.limit("20/minute")
async def download_apk(request: Request, data: dict[str, Any]):
    """提交下载任务."""
    from backend.core.http_client import validate_url
    url = data.get("url", "").strip()
    package = data.get("package", "").strip()
    version = data.get("version", "latest")
    arch = data.get("arch", "unknown")
    detail_url = data.get("detail_url", "").strip()

    if not url or not package:
        raise HTTPException(400, "url and package are required")

    # 包名格式验证
    package = _validate_package_name(package)

    # SSRF 防护: 验证下载 URL
    try:
        url = validate_url(url, allow_all_https=True)
    except ValueError as e:
        raise HTTPException(400, f"URL 无效: {e}")

    settings = get_settings()
    save_path = _safe_save_path(settings, package, version, arch)

    task = DlTask(
        id=f"dl_{uuid.uuid4().hex[:8]}",
        url=url,
        package_name=package,
        version=version,
        arch=arch,
        save_path=save_path,
        detail_url=detail_url,
    )

    manager = get_download_manager()
    await manager.enqueue(task)

    return {"task_id": task.id, "status": "queued"}


@router.post("/download/batch")
@limiter.limit("10/minute")
async def download_batch(request: Request, data: dict[str, Any]):
    """批量提交下载任务."""
    items = data.get("items", [])
    task_ids = []
    settings = get_settings()
    manager = get_download_manager()

    for item in items:
        url = item.get("url", "").strip()
        package = item.get("package", "").strip()
        version = item.get("version", "latest")
        arch = item.get("arch", "unknown")
        detail_url = item.get("detail_url", "").strip()

        if not url or not package:
            continue

        # 包名格式验证
        try:
            package = _validate_package_name(package)
        except HTTPException:
            continue

        save_path = _safe_save_path(settings, package, version, arch)
        task = DlTask(
            id=f"dl_{uuid.uuid4().hex[:8]}",
            url=url,
            package_name=package,
            version=version,
            arch=arch,
            save_path=save_path,
            detail_url=detail_url,
        )
        await manager.enqueue(task)
        task_ids.append(task.id)

    return {"queued": len(task_ids), "task_ids": task_ids}


@router.get("/download/tasks")
async def download_tasks():
    """获取所有下载任务状态."""
    manager = get_download_manager()
    return {"tasks": [t.to_dict() for t in manager.get_all_tasks()]}


@router.post("/download/{task_id}/pause")
async def download_pause(task_id: str):
    manager = get_download_manager()
    if not manager.pause_task(task_id):
        raise HTTPException(404, "任务不存在")
    return {"ok": True}


@router.post("/download/{task_id}/resume")
async def download_resume(task_id: str):
    manager = get_download_manager()
    if not manager.resume_task(task_id):
        raise HTTPException(404, "任务不存在或未暂停")
    return {"ok": True}


@router.post("/download/{task_id}/cancel")
async def download_cancel(task_id: str):
    manager = get_download_manager()
    if not manager.cancel_task(task_id):
        raise HTTPException(404, "任务不存在")
    return {"ok": True}


# ── 记忆化 ─────────────────────────────────────────────────

@router.get("/memo/{package_name}")
async def memo_get(package_name: str):
    """查询包名的历史版本信息."""
    store = get_memo_store()
    result = store.get(package_name)
    if not result:
        return {"found": False}
    return {"found": True, "data": result}


@router.post("/memo")
async def memo_save(data: dict[str, Any]):
    """保存版本信息到记忆."""
    package = data.get("package", "").strip()
    if not package:
        raise HTTPException(400, "package is required")

    store = get_memo_store()
    store.upsert(
        package_name=package,
        version_code=data.get("version_code"),
        version_name=data.get("version_name"),
    )
    return {"ok": True}


@router.get("/memo")
async def memo_list(limit: int = 100):
    """列出所有记忆."""
    store = get_memo_store()
    return {"items": store.list_all(limit)}


# ── 配置 ───────────────────────────────────────────────────

@router.get("/config")
async def config_get():
    settings = get_settings()
    return settings.model_dump(exclude_none=True)


@router.patch("/config")
async def config_update(data: dict[str, Any]):
    settings = get_settings()
    try:
        settings.update(data)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    reload_settings()
    return {"ok": True}


# ── 缓存管理 ───────────────────────────────────────────────

@router.post("/cache/clear")
async def cache_clear():
    """清除所有爬虫结果缓存."""
    cache = get_scraper_cache()
    count = await cache.clear()
    return {"ok": True, "cleared": count}


# ── 每日更新面板 (v2.8.2) ──────────────────────────────────

@router.get("/daily-updates")
async def daily_updates(
    request: Request,
    source: str = Query(None),
    limit: int = Query(20),
):
    """获取 APKPure/APKCombo/APKVision 最近更新游戏列表, 支持条件请求."""
    from backend.cron.update_tracker import get_last_modified
    from backend.db.database import get_connection as _get_conn

    # 条件请求 (RFC 7231)
    last_mod = get_last_modified()
    if_modified_since = request.headers.get("If-Modified-Since")
    if if_modified_since and last_mod:
        try:
            client_time = parsedate_to_datetime(if_modified_since)
            if client_time >= last_mod:
                return Response(status_code=304)
        except (ValueError, TypeError, LookupError):
            pass

    settings = get_settings()

    # v3.4: 每个源使用独立的展示上限
    _display_limits = {
        "apkpure": getattr(settings, "apkpure_display_limit", 60),
        "apkcombo": getattr(settings, "apkcombo_display_limit", 40),
        "apkcombo_trending": getattr(settings, "apkcombo_trending_display_limit", 60),
        "apkvision_updated": getattr(settings, "apkvision_display_limit", 30),
        "apkvision_new": getattr(settings, "apkvision_new_display_limit", 30),
    }

    conn = _get_conn()
    try:
        result: dict = {}
        for src in ("apkpure", "apkcombo", "apkcombo_trending", "apkvision_updated", "apkvision_new"):
            if source and source != src:
                continue
            src_limit = _display_limits.get(src, limit)
            sql = (
                "SELECT app_name, icon_url, detail_url, package_name, "
                "download_count, version_name, updated_at "
                "FROM daily_updates WHERE source = ? "
                "ORDER BY updated_at DESC LIMIT ?"
            )
            rows = conn.execute(sql, (src, src_limit)).fetchall()
            result[src] = [
                {
                    "app_name": r["app_name"],
                    "icon_url": r["icon_url"] or "",
                    "detail_url": r["detail_url"] or "",
                    "package_name": r["package_name"],
                    "download_count": r["download_count"] or "",
                    "version_name": r["version_name"] or "",
                    "updated_at": r["updated_at"],
                }
                for r in rows
            ]
    finally:
        conn.close()

    result["poll_interval"] = getattr(settings, "frontend_poll_interval", 300)
    if last_mod:
        result["last_fetched_at"] = last_mod.strftime("%Y-%m-%d %H:%M:%S")

    headers = {}
    if last_mod:
        headers["Last-Modified"] = format_datetime(last_mod, usegmt=True)
    return JSONResponse(content=result, headers=headers)


@router.post("/daily-updates/refresh")
async def trigger_daily_refresh():
    """全量刷新 — 重新抓取所有数据源的全部页面（最多 45s)."""
    from backend.cron.update_tracker import update_once
    try:
        await asyncio.wait_for(update_once(full_refresh=True), timeout=45.0)
        return {"status": "ok", "message": "全量刷新完成"}
    except asyncio.TimeoutError:
        return {"status": "timeout", "message": "刷新超时，后台继续"}


@router.post("/daily-updates/refresh-incremental")
async def trigger_incremental_refresh():
    """v3.4 增量刷新 — 只抓首页新数据, 速度快且不易被封（最多 30s)."""
    from backend.cron.update_tracker import update_once
    try:
        await asyncio.wait_for(update_once(full_refresh=False), timeout=30.0)
        return {"status": "ok", "message": "增量刷新完成"}
    except asyncio.TimeoutError:
        return {"status": "timeout", "message": "增量刷新超时，后台继续"}


@router.post("/apkpure/unblock")
async def apkpure_unblock(request: Request):
    """v3.3: 手动重置 APKPure 熔断器 (仅允许本地访问).

    APKPure 被封后, 管理员手动换 IP 或等待解封后调用此接口,
    重置熔断器 + 恢复默认更新间隔.
    """
    if request.client and request.client.host != "127.0.0.1":
        raise HTTPException(403, "仅允许本地访问")
    from backend.cron.update_tracker import record_success as _reset
    await _reset("apkpure")
    # 同时恢复默认更新间隔
    settings = get_settings()
    try:
        settings.update({"update_check_interval": 1800})
    except ValueError:
        pass
    logger.info("APKPure 熔断器已被管理员手动重置")
    return {"ok": True, "message": "APKPure 熔断器已重置，更新间隔已恢复为 1800s"}


# ── WebSocket ──────────────────────────────────────────────

_WS_ALLOWED_ORIGINS = {
    "http://127.0.0.1:8000", "http://localhost:8000",
    "http://127.0.0.1:5173", "http://localhost:5173",
    "file://",  # Electron/本地文件
}

def _check_ws_origin(websocket: WebSocket) -> None:
    """v2.8.1: 验证 WebSocket 来源，拒绝不受信任的连接."""
    origin = websocket.headers.get("origin", "")
    if not origin:
        return  # 无来源头 (如本地客户端), 允许
    # file:// 和 Electron 允许任何
    if any(origin.startswith(allowed.replace("://", "://")) for allowed in _WS_ALLOWED_ORIGINS if "://" in allowed):
        return
    if origin == "null" or origin.startswith("file://"):
        return
    logger.warning("WebSocket 来源被拒绝: {}", origin)
    raise HTTPException(403, "不受信任的来源")


@router.websocket("/ws")
async def websocket_global(websocket: WebSocket):
    """全局 WebSocket：接收下载进度、批量任务通知."""
    _check_ws_origin(websocket)
    ws_mgr = get_ws_manager()
    await ws_mgr.connect(websocket)
    try:
        while True:
            _ = await websocket.receive_text()
    except WebSocketDisconnect:
        ws_mgr.disconnect(websocket)


@router.websocket("/ws/{task_id}")
async def websocket_task(websocket: WebSocket, task_id: str):
    """任务专用 WebSocket：订阅特定批量任务的进度."""
    _check_ws_origin(websocket)
    ws_mgr = get_ws_manager()
    await ws_mgr.connect(websocket, task_id=task_id)
    try:
        while True:
            _ = await websocket.receive_text()
    except WebSocketDisconnect:
        ws_mgr.disconnect(websocket, task_id=task_id)


# ── 下载链接提取 ───────────────────────────────────────────

@router.post("/extract-links")
@limiter.limit("20/minute")
async def extract_links(request: Request, data: dict[str, Any]):
    """从指定源的详情页提取下载链接."""
    from backend.core.http_client import validate_url
    source = data.get("source", "").strip()
    detail_url = data.get("detail_url", "").strip()
    package = data.get("package", "").strip()
    version = data.get("version", "").strip()

    if not source or not detail_url:
        raise HTTPException(400, "source and detail_url are required")

    # SSRF 防护: 验证详情页 URL
    try:
        detail_url = validate_url(detail_url)
    except ValueError as e:
        raise HTTPException(400, f"URL 无效: {e}")

    from backend.download.extractors import get_download_page_url
    variants = await extract_download_links(source, detail_url, package=package, version=version)
    best = pick_best_variant(variants)
    download_page_url = get_download_page_url(source, detail_url, package, version)

    return {
        "variants": [
            {"url": v.url, "arch": v.arch, "size": v.size, "source": v.source}
            for v in variants
        ],
        "best": {"url": best.url, "arch": best.arch, "source": best.source} if best else None,
        "download_page_url": download_page_url,
    }
